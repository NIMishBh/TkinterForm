from openpyxl import *
from tkinter import *
import os
from tkinter import Canvas

canvas=Canvas(width=700,height=600)
canvas.create_rectangle(50,50,650,550,outline='grey',fill='grey')
canvas.pack()

current_working_directory=os.getcwd()
excel_location=os.path.join(current_working_directory,'Book1.xlsx')
#wb = load_workbook('C:\\Users\\Nim_Ish\\Desktop\\Book1.xlsx')
wb = load_workbook(excel_location)

sheet = wb.active


def excel():
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 40
    sheet.column_dimensions['D'].width = 50

    sheet.cell(row=1, column=1).value = "Name"
    sheet.cell(row=1, column=2).value = "Contact Nmber"
    sheet.cell(row=1, column=3).value = "Email id"
    sheet.cell(row=1, column=4).value = "Address"


def clear():
    name_field.delete(0, END)
    contact_no_field.delete(0, END)
    email_id_field.delete(0, END)
    address_field.delete(0, END)


def insert():
    if (name_field.get() == "" and
            contact_no_field.get() == "" and
            email_id_field.get() == "" and
            address_field.get() == ""):

        print("empty input")

    else:

        current_row = sheet.max_row
        current_column = sheet.max_column

        sheet.cell(row=current_row + 1, column=1).value = name_field.get()
        sheet.cell(row=current_row + 1, column=2).value = contact_no_field.get()
        sheet.cell(row=current_row + 1, column=3).value = email_id_field.get()
        sheet.cell(row=current_row + 1, column=4).value = address_field.get()

        wb.save('C:\\Users\\Nim_Ish\\Desktop\\Book1.xlsx')
        name_field.focus_set()
        clear()


if __name__ == "__main__":
    #root = Tk()
    #canvas.minsize(700,700)
    #canvas.configure(background='grey')
    #canvas.title("registration form")
    #root.geometry("500x300")

    excel()
    heading = Label(canvas, text="Form",bg='SeaGreen1',font=('Verdana', 15),padx=5,pady=2)
    name = Label(canvas, text="Name", bg='SeaGreen1',padx=5,pady=2)
    contact_no = Label(canvas, text="Contact No.",bg='SeaGreen1',padx=5,pady=2)
    email_id = Label(canvas, text="Email id", bg='SeaGreen1',padx=5,pady=2)
    address = Label(canvas, text="Address",bg='SeaGreen1',padx=5,pady=2)

    heading.place(relx=0.5,rely=0.2,anchor='center')
    name.place(relx=0.2,rely=0.4)
    contact_no.place(relx=0.2,rely=0.5)
    email_id.place(relx=0.2,rely=0.6)
    address.place(relx=0.2,rely=0.7)

    name_field = Entry(canvas)
    contact_no_field = Entry(canvas)
    email_id_field = Entry(canvas)
    address_field = Entry(canvas)

    name_field.place(relx=0.6,rely=0.4)
    contact_no_field.place(relx=0.6,rely=0.5)
    email_id_field.place(relx=0.6,rely=0.6)
    address_field.place(relx=0.6,rely=0.7)

    excel()

    submit = Button(canvas, text="Submit", fg="Black"
                    , command=insert,bg='SeaGreen1',font=('Verdana', 10),padx=5,pady=2)
    submit.place(relx=0.5,rely=0.85,anchor='center')

     #start the GUI
    canvas.mainloop()
